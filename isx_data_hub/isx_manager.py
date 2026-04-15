import sqlite3
import pandas as pd
import os

class ISXManager:
    DB_NAME = "isx_nucleus.db"
    EXCEL_NAME = "isx_nucleus.xlsx"
    
    def __init__(self, output_dir="."):
        self.output_dir = output_dir
        self.db_path = os.path.join(output_dir, self.DB_NAME)
        self.excel_path = os.path.join(output_dir, self.EXCEL_NAME)
        self._init_db()

    def _init_db(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Create companies table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                symbol TEXT UNIQUE,
                name TEXT,
                sector TEXT
            )
        ''')
        
        # Create prices table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                date DATE,
                open REAL,
                high REAL,
                low REAL,
                close REAL,
                volume INTEGER,
                amount REAL,
                UNIQUE(company_id, date),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        ''')
        
        # Create news table for RAG
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS news (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                date DATE,
                headline TEXT,
                content TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        ''')
        
        # Create stakeholders table for RAG
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stakeholders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                owner_name TEXT,
                share_percentage REAL,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        ''')
        
        # Create company_status table for RAG
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS company_status (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                status TEXT,
                update_date DATE,
                details TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        ''')
        
        conn.commit()
        conn.close()

    def save_company(self, symbol, name, sector="Unknown"):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('INSERT OR IGNORE INTO companies (symbol, name, sector) VALUES (?, ?, ?)', (symbol, name, sector))
            conn.commit()
            cursor.execute('SELECT id FROM companies WHERE symbol = ?', (symbol,))
            company_id = cursor.fetchone()[0]
            return company_id
        finally:
            conn.close()

    def save_prices(self, company_id, df):
        if df is None or df.empty:
            return
        
        conn = sqlite3.connect(self.db_path)
        try:
            for _, row in df.iterrows():
                conn.execute('''
                    INSERT OR REPLACE INTO prices 
                    (company_id, date, open, high, low, close, volume, amount) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (company_id, row['date'].strftime('%Y-%m-%d'), row['open'], row['high'], row['low'], row['close'], row['volume'], row['amount']))
            conn.commit()
        finally:
            conn.close()

    def export_to_excel(self):
        conn = sqlite3.connect(self.db_path)
        try:
            # Join companies and prices
            query = '''
                SELECT c.symbol, c.name, c.sector, p.date, p.open, p.high, p.low, p.close, p.volume, p.amount
                FROM prices p
                JOIN companies c ON p.company_id = c.id
                ORDER BY c.symbol, p.date DESC
            '''
            df = pd.read_sql_query(query, conn)
            
            # Save to Excel with openpyxl engine
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='All Data', index=False)
                
                # Access the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['All Data']
                
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter

                # Format Headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill

                # Format Numbers (Volume and Amount with commas)
                for col_idx in [9, 10]:  # volume(I) and amount(J) are 9th and 10th cols
                    letter = get_column_letter(col_idx)
                    for cell in worksheet[letter]:
                        if cell.row > 1:
                            cell.number_format = '#,##0'

                # Auto-adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
                    
            print(f"Exported clean formatted Excel to {self.excel_path}")
        finally:
            conn.close()

if __name__ == "__main__":
    manager = ISXManager()
    print("Database initialized.")
